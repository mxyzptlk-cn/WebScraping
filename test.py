#!/usr/bin/env python3
# -*- coding:utf-8 -*- 
# Author: Mxyzptlk
# Date: 2018-09-18




"""
 It's not a bug, it's a feature.
             ┏┓   ┏┓
            ┏┛┻━━━┛┻┓
            ┃   ☃   ┃
            ┃ ┳┛ ┗┳ ┃
            ┃   ┻   ┃
            ┗━┓   ┏━┛
              ┃   ┗━━━┓
              ┃神兽保佑 ┣┓
              ┃永无BUG ┏┛
              ┗┓┓┏━┳┓┏┛
               ┃┫┫  ┃┫┫
               ┗┻┛  ┗┻┛
"""


# 1	医生
# 2	护士
# 3	医技
# 4	药房
# 5	后勤
# 6	物资
# 7	收费员

import os, time
import pyautogui as ag
import pyperclip as pc
from time import sleep
from openpyxl import load_workbook


def get_list():
    data_sheet = load_workbook('通讯录.xlsx')
    data = data_sheet.active
    data_list = []
    for i in range(2, 500):
        if data['A' + str(i)].value is None:
            break
        else:
            data_item = []
            data_item.append(data['A' + str(i)].value)
            data_item.append(data['B' + str(i)].value)
            data_item.append(data['C' + str(i)].value)
            data_item.append(data['D' + str(i)].value)
            data_list.append(data_item)
    return data_list


data_sheet = load_workbook('通讯录.xlsx')
data = data_sheet.active

departmet = None

for _ in range(6,328):
    if data['a'+str(_)].value:
        departmet = data['a'+str(_)].value
    data['e' + str(_)] = data['b'+str(_)].value + '-' + departmet


data_sheet.save('test2.xlsx')

